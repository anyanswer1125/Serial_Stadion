import os
import sys
from datetime import datetime
from openpyxl import load_workbook, Workbook
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QVBoxLayout, QLineEdit, QPushButton, QLabel, QTextEdit, QFileDialog, QWidget, QInputDialog
)
from PySide6.QtCore import QFileSystemWatcher
from PySide6.QtCore import QEvent
from PySide6.QtWidgets import QMessageBox

DEFAULT_EXCEL_FILE = "data.xlsm"  # 기본 데이터베이스 파일 이름
QSS_FILE = "style.qss"  # QSS 파일 경로


class BarcodeApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("바코드 처리 프로그램")
        self.current_file = DEFAULT_EXCEL_FILE  # 현재 데이터베이스 파일
        
        
        self.watcher = QFileSystemWatcher([self.current_file])
        self.watcher.fileChanged.connect(self.on_file_changed)

        # 중앙 위젯 설정
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        self.layout = QVBoxLayout()

        # 바코드 입력창
        self.input_line = QLineEdit()
        self.input_line.setPlaceholderText("바코드를 입력하세요")
        self.layout.addWidget(self.input_line)

        # 처리 버튼
        self.button = QPushButton("처리")
        self.button.clicked.connect(self.on_process_clicked)
        self.layout.addWidget(self.button)

        # 결과 표시
        self.result_label = QLabel("")
        self.layout.addWidget(self.result_label)

        # 최근 항목 표시
        self.recent_label = QLabel("최근 항목")
        self.layout.addWidget(self.recent_label)

        self.recent_items_text = QTextEdit()
        self.recent_items_text.setReadOnly(True)
        self.layout.addWidget(self.recent_items_text)

        central_widget.setLayout(self.layout)

        # 메뉴바 추가
        self.create_menu_bar()

        self.update_recent_items()
        self.activateWindow()
        self.input_line.setFocus()
        self.input_line.returnPressed.connect(self.on_process_clicked)


    def create_menu_bar(self):
        """메뉴바 생성"""
        menu_bar = self.menuBar()

        # 파일 메뉴
        file_menu = menu_bar.addMenu("파일")

        # "다른 이름으로 저장하기" 메뉴
        save_action = file_menu.addAction("다른 이름으로 저장하기")
        save_action.triggered.connect(self.save_as)

        # "불러오기" 메뉴
        load_action = file_menu.addAction("불러오기")
        load_action.triggered.connect(self.load_file)

    def update_file_name_label(self):
        """현재 파일명을 메뉴바 옆에 업데이트"""
        self.file_name_label.setText(f"현재 파일: {os.path.basename(self.current_file)}")

    def on_process_clicked(self):
        barcode = self.input_line.text().strip()
        if not barcode:
            self.result_label.setText("바코드를 입력해주세요.")
            return

        msg = process_barcode(barcode, self.current_file, self)
        self.result_label.setText(msg)
        self.input_line.clear()
        self.update_recent_items()


    def save_as(self):
        """다른 이름으로 파일 저장"""
        file_path, _ = QFileDialog.getSaveFileName(self, "파일 저장", "", "Excel Files (*.xlsm *.xlsx)")
        if file_path:
            try:
                wb = load_workbook(self.current_file, keep_vba=True)
                wb.save(file_path)
                wb.close()
                self.result_label.setText(f"파일이 저장되었습니다: {file_path}")
            except Exception as e:
                self.result_label.setText(f"파일 저장 중 오류 발생: {str(e)}")

    def load_file(self):
        """다른 파일 불러오기"""
        file_path, _ = QFileDialog.getOpenFileName(self, "파일 열기", "", "Excel Files (*.xlsm *.xlsx)")
        if file_path:
            self.current_file = file_path
            self.watcher.removePaths(self.watcher.files())  # 기존 감시자 제거
            self.watcher.addPath(file_path)  # 새 파일 감시자 등록
            self.update_recent_items()
            self.result_label.setText(f"파일이 변경되었습니다: {file_path}")


    def update_recent_items(self):
        """최근 항목 업데이트"""
        recent_items = get_recent_items(self.current_file)
        self.recent_items_text.setText(recent_items)
    def prompt_max_duplicate(self):
    # 최대 중복 횟수를 입력받는 팝업
        max_duplicate, ok = QInputDialog.getInt(
            self,                           # 부모 위젯 (여기서는 BarcodeApp 클래스의 인스턴스)
            "최대 중복 횟수 설정",           # 팝업창의 제목 (타이틀)
            "최대 중복 횟수를 입력해주세요:", # 사용자에게 보여줄 메시지 (라벨 텍스트)
            10,                              # 기본값 (기본으로 입력창에 표시되는 값, 여기서는 10)
            1,                              # 최소값 (사용자가 입력할 수 있는 가장 작은 값, 여기서는 1)
            1000,                           # 최대값 (사용자가 입력할 수 있는 가장 큰 값, 여기서는 1000)
            1                               # 스텝 값 (입력 시 숫자가 얼마나 증가/감소하는지, 여기서는 1씩)
        )                                                                                           
        return max_duplicate if ok else None
    def on_file_changed(self):
        # 엑셀 파일이 변경되었을 때 자동으로 최근 항목 업데이트
        self.update_recent_items()
    def changeEvent(self, event):
        """창 활성화 시 바코드 입력창에 포커스를 강제로 설정"""
        if event.type() == QEvent.ActivationChange and self.isActiveWindow():
            self.input_line.setFocus()
        super().changeEvent(event)

def ensure_excel_file(file_path):
    """데이터 관리를 위한 엑셀 파일 생성 확인"""
    try:
        wb = load_workbook(file_path, keep_vba=True)
        ws = wb.active

        # 기존에 F1까지 사용하던 헤더를 새로운 형식으로 변경
        if ws.max_column < 6:
            ws["A1"] = "바코드"
            ws["B1"] = "날짜"
            ws["C1"] = "횟수"
            ws["D1"] = "비고"
            ws["E1"] = "최대중복"
            wb.save(file_path)
        wb.close()
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "바코드"
        ws["B1"] = "날짜"
        ws["C1"] = "횟수"
        ws["D1"] = "비고"
        ws["E1"] = "최대중복"
        wb.save(file_path)

def show_max_duplicate_popup(app_window, barcode, max_duplicate):
    """최대 중복 횟수 초과 시 경고 팝업"""
    msg_box = QMessageBox(app_window)
    msg_box.setIcon(QMessageBox.Warning)
    msg_box.setWindowTitle("최대 중복 횟수 초과")
    msg_box.setText(f"바코드 {barcode}의 최대 중복 횟수 ({max_duplicate})를 초과했습니다.\n더 이상 등록할 수 없습니다.")
    msg_box.setStandardButtons(QMessageBox.Ok)
    msg_box.exec()

def process_barcode(barcode, file_path, app_window):
    """바코드를 엑셀 데이터베이스에 추가하고 처리"""
    wb = load_workbook(file_path, keep_vba=True)
    ws = wb.active

    # 바코드가 이미 있는지 확인
    existing_barcodes = [ws.cell(row=row, column=1).value for row in range(2, ws.max_row + 1)]
    barcode_exists = barcode in existing_barcodes

    # 현재 시각 (날짜 컬럼에 들어갈 값)
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M')

    if not barcode_exists:
        max_duplicate = app_window.prompt_max_duplicate()
        if max_duplicate is None:
            wb.close()
            return "등록이 취소되었습니다."

        # 신규 바코드 첫 번째 등록
        new_row = ws.max_row + 1
        ws.cell(row=new_row, column=1, value=barcode)
        ws.cell(row=new_row, column=2, value=now_str)  # 날짜
        ws.cell(row=new_row, column=3, value="1 회")  # 횟수
        ws.cell(row=new_row, column=4, value="신규 등록")  # 비고
        ws.cell(row=new_row, column=5, value=f"{max_duplicate}회")  # 최대중복

        # ✅ 입력창 초기화
        app_window.input_line.clear()
        app_window.input_line.setFocus()
    else:
        # 이미 존재하는 경우 중복 횟수 업데이트
        max_row = ws.max_row + 1
        count = sum(1 for row in range(2, max_row) if ws.cell(row=row, column=1).value == barcode)
        max_duplicate = next(
            (ws.cell(row=row, column=5).value for row in range(2, max_row) if ws.cell(row=row, column=1).value == barcode),
            "1회"
        )

        # 최대 중복 횟수 초과 시 등록 차단
        if count >= int(max_duplicate.replace("회", "")):
            wb.close()
            show_max_duplicate_popup(app_window, barcode, max_duplicate)
            app_window.input_line.clear()
            app_window.input_line.setFocus()
            return f"등록 불가: 바코드 {barcode}의 최대 중복 횟수({max_duplicate})를 초과했습니다."

        # 상태 업데이트
        state = "최대 한도 도달" if count + 1 >= int(max_duplicate.replace("회", "")) else "중복 사용"

        # 데이터 추가
        ws.cell(row=max_row, column=1, value=barcode)
        ws.cell(row=max_row, column=2, value=now_str)  # 날짜
        ws.cell(row=max_row, column=3, value=f"{count + 1} 회")  # 횟수
        ws.cell(row=max_row, column=4, value=state)  # 비고
        ws.cell(row=max_row, column=5, value=max_duplicate)  # 최대중복

    wb.save(file_path)
    wb.close()

    # ✅ 바코드 입력창 초기화 및 포커스
    app_window.input_line.clear()
    app_window.input_line.setFocus()

    return f"바코드 {barcode} 처리 완료 (중복: {count + 1}/{max_duplicate})"

def get_recent_items(file_path, limit=10):
    """최근 항목 가져오기 (일정한 간격 유지)"""
    wb = load_workbook(file_path, keep_vba=True)
    ws = wb.active

    max_row = ws.max_row
    recent_items = []

    # 칼럼 
    # 너비 설정 (문자열 길이를 균일하게 유지)
    COL_WIDTH = {
        "barcode": 22,   # 바코드 길이 고정
        "timestamp": 18, # YYYY-MM-DD HH:MM
        "count": 12,      # "99 회"
        "status": 20,    # "최대 사용 횟수 도달"
        "max_count": 6   # "99회"
    }

    for row in range(max(max_row - limit + 1, 2), max_row + 1):  # 마지막 `limit`개만 가져옴
        barcode = str(ws.cell(row=row, column=1).value).ljust(COL_WIDTH["barcode"])  # 왼쪽 정렬
        timestamp = str(ws.cell(row=row, column=2).value).rjust(COL_WIDTH["timestamp"])
        duplicate_count = str(ws.cell(row=row, column=3).value).rjust(COL_WIDTH["count"])  # 오른쪽 정렬
        status = str(ws.cell(row=row, column=4).value).rjust(COL_WIDTH["status"])  # 왼쪽 정렬
        max_count = str(ws.cell(row=row, column=5).value).rjust(COL_WIDTH["max_count"])  # 오른쪽 정렬

        # recent_items.append(f"{barcode} {timestamp} {duplicate_count} {status} {max_count}") #백업
        recent_items.append(f"   {barcode} {timestamp} {duplicate_count} {status}") 

    wb.close()
    return "\n".join(recent_items)


def load_qss(app, qss_file):
    """QSS 파일을 로드하여 스타일 적용"""
    try:
        with open(qss_file, "r", encoding="utf-8") as f:
            app.setStyleSheet(f.read())
    except FileNotFoundError:
        print(f"QSS 파일을 찾을 수 없습니다: {qss_file}")


if __name__ == "__main__":
    ensure_excel_file(DEFAULT_EXCEL_FILE)
    app = QApplication(sys.argv)

    # QSS 로드 및 핫 리로드 설정
    load_qss(app, QSS_FILE)
    watcher = QFileSystemWatcher([QSS_FILE])
    watcher.fileChanged.connect(lambda: load_qss(app, QSS_FILE))  # QSS 변경 시 자동 적용

    window = BarcodeApp()
    window.show()
    sys.exit(app.exec())
