import os
import sys
from datetime import datetime
from openpyxl import load_workbook, Workbook
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QVBoxLayout, QHBoxLayout, QLineEdit, QPushButton, QLabel,
    QTextEdit, QFileDialog, QWidget, QTableWidget, QTableWidgetItem, QMessageBox, QScrollBar
    ,QComboBox,QSizePolicy, QHeaderView
)
from PySide6.QtCore import Qt, QEvent

DEFAULT_EXCEL_FILE = "data.xlsm"  # 기본 데이터 파일


class BarcodeApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("할인쿠폰R0.2_preview")
        self.resize(900, 600)  # 가로 크기 확대
        

        self.current_file = DEFAULT_EXCEL_FILE

        # ✅ 파일 존재 여부 확인 (파일이 없으면 경고창 출력 후 종료)
        if not os.path.exists(self.current_file):
            QMessageBox.critical(self, "파일 없음", f"데이터 파일 '{self.current_file}'이(가) 없습니다.\n파일을 확인하세요.")
            sys.exit(1)  # 프로그램 종료



        # 중앙 위젯
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        # 레이아웃 설정
        main_layout = QHBoxLayout()  # 가로 레이아웃
        left_layout = QVBoxLayout()  # 바코드 입력 및 최근 항목
        right_layout = QVBoxLayout()  # 검색 결과

        # 바코드 입력창 (검색 & 처리 통합)
        self.input_line = QLineEdit()
        self.input_line.setPlaceholderText("바코드를 입력 후 엔터 (처리 & 검색)")
        self.input_line.returnPressed.connect(self.on_process_and_search)
        self.input_line.setFixedHeight(100)
        self.input_line.setStyleSheet("font-size: 20px;")
        
        left_layout.addWidget(self.input_line)
        

        # "검색(만)" 버튼 클릭 시 검색 기능만 수행하도록 설정
        self.button = QPushButton("검색(만)")
        self.button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)  # 가로 크기 자동 확장
        self.button.setFixedHeight(50)
        self.button.setStyleSheet("font-size:20px")
        self.button.clicked.connect(self.on_search_only)


        # 최대 중복 선택 드롭다운
        self.max_duplicate_selector = QComboBox()
        self.max_duplicate_selector.addItems(["5", "10", "20"])  # 선택지 추가
        self.max_duplicate_selector.setCurrentText("10")  # 기본값 10
        self.max_duplicate_selector.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)  # 가로 크기 자동 확장
        self.max_duplicate_selector.setFixedHeight(50)
        self.max_duplicate_selector.setStyleSheet("font-size: 20px")

        # 버튼 & 드롭다운 정렬을 위한 가로 레이아웃
        button_layout = QHBoxLayout()
        button_layout.addWidget(self.button, 8)  # 비율 8
        button_layout.addWidget(self.max_duplicate_selector, 2)  # 비율 2
        button_layout.setStretch(0, 8)  # 버튼 비율 8
        button_layout.setStretch(1, 2)  # 드롭다운 비율 2

        # 배치
        left_layout.addLayout(button_layout)

        # 최근 항목 테이블로 변경
        self.recent_table = QTableWidget()
        self.recent_table.setColumnCount(4)  # 4개 컬럼 (바코드, 날짜, 횟수, 비고)
        self.recent_table.setHorizontalHeaderLabels(["바코드", "날짜", "시간", "횟수"])

        # 가변 크기 설정
        # self.recent_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)  # 테이블 크기 자동 조정
        self.recent_table.setColumnWidth(0,150)
        self.recent_table.setColumnWidth(2,50)
        self.recent_table.setEditTriggers(QTableWidget.NoEditTriggers)  # 편집 불가
        # self.recent_table.setSelectionBehavior(QTableWidget.SelectRows)  # 행 전체 선택
        # self.recent_table.setSelectionMode(QTableWidget.SingleSelection)  # 단일 선택
        left_layout.addWidget(self.recent_table)  # 왼쪽 레이아웃에 추가


        # 검색 결과 테이블
        self.search_table = QTableWidget()
        self.search_table.setColumnCount(4)  # 3개 데이터 + 삭제 버튼
        self.search_table.setHorizontalHeaderLabels(["날짜", "횟수", "비고", "삭제"])
        self.search_table.setColumnWidth(0, 150)  # 날짜
        self.search_table.setColumnWidth(1, 60)  # 횟수
        self.search_table.setColumnWidth(2, 80)  # 비고
        self.search_table.setColumnWidth(3, 30)  # 삭제 버튼
        right_layout.addWidget(self.search_table)

        # 레이아웃 구성
        main_layout.addLayout(left_layout, 1)  # 왼쪽 1 비율
        main_layout.addLayout(right_layout, 1)  # 오른쪽 2 비율

        central_widget.setLayout(main_layout)

        self.update_recent_items()

    def on_process_and_search(self):
        """바코드를 처리하고 동시에 검색"""
        barcode = self.input_line.text().strip()
        if not barcode:
            QMessageBox.warning(self, "경고", "바코드를 입력하세요.")
            return

        # 선택한 최대 중복 횟수 가져오기
        max_duplicate = int(self.max_duplicate_selector.currentText())

        # 바코드 처리 (등록)
        msg = process_barcode(barcode, self.current_file, max_duplicate)

        # UI 업데이트
        self.input_line.clear()
        self.update_recent_items(scroll_to_bottom=True)  # ✅ 최근 항목 맨 아래로 자동 스크롤
        self.perform_search(barcode)  # ✅ 검색 결과 자동 표시


    def update_recent_items(self, scroll_to_bottom=False):
        """최근 항목을 테이블(QTableWidget) 형식으로 업데이트"""
        recent_items = get_recent_items(self.current_file, limit=100)  # 최근 100개 가져오기
        recent_items = [line.split() for line in recent_items.split("\n") if line]  # 데이터 가공

        self.recent_table.setRowCount(len(recent_items))  # 행 개수 설정

        for row, item in enumerate(recent_items):
            for col, data in enumerate(item):
                self.recent_table.setItem(row, col, QTableWidgetItem(data))  # 테이블에 데이터 삽입

        if scroll_to_bottom:
            self.recent_table.scrollToBottom()  # ✅ 최신 항목이 가장 아래로 가도록 스크롤 이동

    def perform_search(self, barcode):
        """입력된 바코드를 검색하여 테이블에 표시"""
        result = find_barcode_in_excel(self.current_file, barcode)

        if result:
            self.update_search_results(result, barcode)
        else:
            QMessageBox.information(self, "검색 결과", "해당 바코드가 없습니다.")

    def update_search_results(self, results, barcode):
        """검색 결과 테이블을 업데이트하고 '최대 중복' 컬럼 추가"""
        self.search_table.setColumnCount(4)  # ✅ 컬럼 5개로 변경
        self.search_table.setHorizontalHeaderLabels(["날짜", "횟수", "비고", "삭제"])

        self.search_table.setRowCount(len(results))

        for row, (date, count, remark, max_dup) in enumerate(results):
            self.search_table.setItem(row, 0, QTableWidgetItem(date))
            self.search_table.setItem(row, 1, QTableWidgetItem(count))
            self.search_table.setItem(row, 2, QTableWidgetItem(remark))
            # self.search_table.setItem(row, 3, QTableWidgetItem(max_dup))  # ✅ 최대 중복 컬럼 추가

            # 삭제 버튼 추가
            delete_button = QPushButton("X")
            delete_button.setStyleSheet("border: 0px ; border-radius: 5px;font-size: 20px; font-weight: 1000;color: white; background-color: rgb(255, 60, 60);")
            delete_button.clicked.connect(lambda _, r=row: self.delete_barcode_entry(barcode, r))
            self.search_table.setCellWidget(row, 3, delete_button)  # ✅ 5번째 컬럼 (삭제 버튼)

    def delete_barcode_entry(self, barcode, row):
        """선택한 바코드 행을 삭제하기 전에 두 번의 확인 메시지를 표시"""
        date = self.search_table.item(row, 0).text()
        count = self.search_table.item(row, 1).text()
    
        # 첫 번째 확인 메시지
        first_confirm = QMessageBox.question(
            self, "삭제 확인", "삭제한 후에는 되돌릴 수 없습니다. 이를 확인했습니까?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
    
        if first_confirm != QMessageBox.Yes:
            return  # 사용자가 '아니요'를 선택하면 삭제 취소
    
        # 두 번째 확인 메시지
        second_confirm = QMessageBox.question(
            self, "최종 삭제 확인", "정말 삭제하시겠습니까?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
    
        if second_confirm != QMessageBox.Yes:
            return  # 사용자가 '아니요'를 선택하면 삭제 취소
    
        # 최종 삭제 진행
        if delete_barcode_from_excel(self.current_file, barcode, date, count):
            QMessageBox.information(self, "삭제 완료", "바코드 항목이 삭제되었습니다.")
            self.perform_search(barcode)  # ✅ 검색 결과 갱신
            self.update_recent_items(scroll_to_bottom=True)  # ✅ 삭제 후 최근 항목도 갱신
        else:
            QMessageBox.warning(self, "삭제 실패", "삭제할 수 없습니다.")
    
    def changeEvent(self, event):
        """창 활성화 시 바코드 입력창에 포커스 강제 설정"""
        if event.type() == QEvent.ActivationChange and self.isActiveWindow():
            self.input_line.setFocus()
        super().changeEvent(event)

    def on_search_only(self):
        """검색(만) 버튼 클릭 시 바코드를 검색"""
        barcode = self.input_line.text().strip()
        if not barcode:
            QMessageBox.warning(self, "경고", "바코드를 입력하세요.")
            return

        self.perform_search(barcode)  # ✅ 검색 수행


def process_barcode(barcode, file_path, max_duplicate):
    """바코드를 엑셀에 추가하면서 최대 중복 횟수를 확인"""
    wb = load_workbook(file_path, keep_vba=True)
    ws = wb.active

    now_str = datetime.now().strftime('%Y-%m-%d %H:%M')

    # 바코드 중복 횟수 계산
    existing_rows = [row for row in range(2, ws.max_row + 1) if str(ws.cell(row, 1).value) == barcode]
    count = len(existing_rows) + 1  # 현재 바코드 사용 횟수

    # 기존에 최대 중복 횟수가 등록되어 있으면 가져오고, 없으면 UI에서 선택한 값으로 설정
    if existing_rows:
        max_duplicate_cell = ws.cell(existing_rows[0], 5)  # 5번째 컬럼 (최대 중복)
        if max_duplicate_cell.value:
            max_duplicate = int(str(max_duplicate_cell.value).replace("회", "").strip())

    # 중복 횟수가 최대 중복 횟수에 도달하면 팝업 표시
    if count == max_duplicate:
        QMessageBox.information(None, "사용 완료", f"바코드 {barcode}의 최대 중복 횟수({max_duplicate})에 도달했습니다.")

    # 중복 횟수가 최대 중복 횟수를 초과하면 등록 안 됨
    if count > max_duplicate:
        QMessageBox.warning(None, "등록 불가", f"바코드 {barcode}의 최대 중복 횟수({max_duplicate})를 초과했습니다.")
        wb.close()
        return f"등록 불가: 최대 중복 횟수({max_duplicate}) 초과"

    # 신규 등록이면 최대 중복 값도 함께 저장
    if count == 1:
        remark = "신규 등록"
        max_dup_value = f"{max_duplicate}회"
    elif count == max_duplicate:
        remark = f"{max_duplicate}회 완료"  # ✅ 최대 횟수 도달 시 "완료" 처리
        max_dup_value = ws.cell(existing_rows[0], 5).value
    else:
        remark = "중복 사용"
        max_dup_value = ws.cell(existing_rows[0], 5).value

    # ✅ 데이터 추가
    ws.append([barcode, now_str, f"{count} 회", remark, max_dup_value])

    # ✅ 저장 순서 수정
    wb.save(file_path)
    wb.close()

    return f"바코드 {barcode} 처리 완료 ({count}/{max_duplicate})"


def find_barcode_in_excel(file_path, barcode):
    """엑셀에서 바코드 검색하여 최대 중복 횟수까지 가져오기"""
    wb = load_workbook(file_path, keep_vba=True)
    ws = wb.active

    results = [
        (ws.cell(row, 2).value, ws.cell(row, 3).value, ws.cell(row, 4).value, ws.cell(row, 5).value)  # 최대 중복 포함
        for row in range(2, ws.max_row + 1)
        if str(ws.cell(row, 1).value) == barcode
    ]

    wb.close()
    return results if results else None


def delete_barcode_from_excel(file_path, barcode, date, count):
    """엑셀에서 특정 바코드 행 삭제"""
    wb = load_workbook(file_path, keep_vba=True)
    ws = wb.active

    rows_to_keep = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if not (str(row[0].value) == barcode and str(row[1].value) == date and str(row[2].value) == count):
            rows_to_keep.append([cell.value for cell in row])

    ws.delete_rows(2, ws.max_row)
    
    for row_data in rows_to_keep:
        ws.append(row_data)

    wb.save(file_path)
    wb.close()

    return True


def get_recent_items(file_path, limit=100):
    """최근 100개 항목 가져오기"""
    wb = load_workbook(file_path, keep_vba=True)
    ws = wb.active

    max_row = ws.max_row
    recent_items = "\n".join(
        f"{ws.cell(row, 1).value} {ws.cell(row, 2).value} {ws.cell(row, 3).value} {ws.cell(row, 4).value}"
        for row in range(max(max_row - limit, 1), max_row + 1)
    )

    wb.close()
    return recent_items


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = BarcodeApp()
    window.show()
    sys.exit(app.exec())
